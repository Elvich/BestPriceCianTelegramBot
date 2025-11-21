"""
Модуль для экспорта данных из базы в Excel формат.
Создает Excel файлы с данными о квартирах.
"""

import pandas as pd
import os
import sys
from datetime import datetime
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Добавляем путь к родительской директории
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from DB.apartment_service import ApartmentService
from DB.reaction_service import ReactionService
from DB.Models import Apartment


class ExcelExporter:
    """Класс для экспорта данных в Excel формат"""
    
    @staticmethod
    async def export_apartments_to_excel(
        filename: Optional[str] = None,
        max_price: Optional[int] = None,
        min_price: Optional[int] = None,
        metro_stations: Optional[List[str]] = None,
        limit: Optional[int] = None
    ) -> str:
        """
        Экспортирует объявления о квартирах в Excel файл
        
        Args:
            filename: Имя файла (если не указано, генерируется автоматически)
            max_price: Максимальная цена
            min_price: Минимальная цена
            metro_stations: Список станций метро
            limit: Ограничение количества записей (None = все записи)
            
        Returns:
            str: Путь к созданному файлу
        """
        
        # Получаем данные из БД
        apartments = await ApartmentService.get_apartments(
            limit=limit,
            min_price=min_price,
            max_price=max_price,
            metro_stations=metro_stations,
            only_active=True
        )
        
        if not apartments:
            raise ValueError("Нет данных для экспорта")
        
        # Генерируем имя файла если не указано
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"apartments_export_{timestamp}.xlsx"
        
        # Создаем DataFrame
        data = []
        for apt in apartments:
            # Собираем информацию о станциях метро
            metro_info = []
            for metro in apt.metro_stations:
                metro_info.append(f"{metro.station_name} ({metro.travel_time})")
            metro_str = "; ".join(metro_info) if metro_info else "Не указано"
            
            # Форматируем цены
            price_formatted = f"{apt.price:,} ₽" if apt.price else "Не указана"
            price_per_sqm_formatted = f"{apt.price_per_sqm:,} ₽/м²" if apt.price_per_sqm else "Не указана"
            
            data.append({
                'ID': apt.cian_id,
                'Название': apt.title,
                'Цена': price_formatted,
                'Цена за м²': price_per_sqm_formatted,
                'Цена (число)': apt.price or 0,
                'Цена за м² (число)': apt.price_per_sqm or 0,
                'Адрес': apt.address or "Не указан",
                'Станции метро': metro_str,
                'Ссылка': apt.url,
                'Дата добавления': apt.first_seen.strftime("%d.%m.%Y %H:%M"),
                'Последнее обновление': apt.last_updated.strftime("%d.%m.%Y %H:%M"),
                'Активно': "Да" if apt.is_active else "Нет"
            })
        
        df = pd.DataFrame(data)
        
        # Создаем Excel файл с красивым форматированием
        ExcelExporter._create_formatted_excel(df, filename)
        
        return os.path.abspath(filename)
    
    @staticmethod
    async def export_browse_apartments_to_excel(
        user_id: int,
        filename: Optional[str] = None,
        limit: Optional[int] = None
    ) -> str:
        """
        Экспортирует квартиры из раздела "Просмотр квартир" для конкретного пользователя
        
        Args:
            user_id: ID пользователя в Telegram
            filename: Имя файла (если не указано, генерируется автоматически)
            limit: Ограничение количества записей (None = все записи)
            
        Returns:
            str: Путь к созданному файлу
        """
        
        # Получаем квартиры для просмотра (исключаем дизлайкнутые пользователем)
        apartments = await ApartmentService.get_apartments(
            limit=limit,
            only_active=True,
            only_production=True,
            exclude_disliked_for_user=user_id
        )
        
        if not apartments:
            raise ValueError("Нет квартир для экспорта")
        
        # Генерируем имя файла
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"browse_apartments_{timestamp}.xlsx"
        
        # Подготавливаем данные
        data = await ExcelExporter._prepare_apartments_data(apartments, user_id)
        df = pd.DataFrame(data)
        
        # Создаем Excel файл
        ExcelExporter._create_formatted_excel(
            df, filename, 
            sheet_title="Просмотр квартир",
            info_title="Квартиры для просмотра",
            description="Все доступные квартиры (исключены ваши дизлайки)"
        )
        
        return os.path.abspath(filename)
    
    @staticmethod
    async def export_user_liked_apartments_to_excel(
        user_id: int,
        filename: Optional[str] = None,
        limit: Optional[int] = None
    ) -> str:
        """
        Экспортирует лайкнутые пользователем квартиры
        
        Args:
            user_id: ID пользователя в Telegram
            filename: Имя файла (если не указано, генерируется автоматически)
            limit: Ограничение количества записей (None = все записи)
            
        Returns:
            str: Путь к созданному файлу
        """
        
        # Получаем лайкнутые квартиры
        apartments = await ReactionService.get_user_liked_apartments(user_id, limit)
        
        if not apartments:
            raise ValueError("У вас пока нет лайкнутых квартир")
        
        # Генерируем имя файла
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"liked_apartments_{timestamp}.xlsx"
        
        # Подготавливаем данные
        data = await ExcelExporter._prepare_apartments_data(apartments, user_id)
        df = pd.DataFrame(data)
        
        # Создаем Excel файл
        ExcelExporter._create_formatted_excel(
            df, filename,
            sheet_title="Мои лайки",
            info_title="Ваши лайкнутые квартиры",
            description="Квартиры, которые вы добавили в избранное"
        )
        
        return os.path.abspath(filename)
    
    @staticmethod
    async def _prepare_apartments_data(apartments: List[Apartment], user_id: int) -> List[dict]:
        """
        Подготавливает данные квартир для экспорта с учетом реакций пользователя
        
        Args:
            apartments: Список квартир
            user_id: ID пользователя для получения реакций
            
        Returns:
            List[dict]: Подготовленные данные
        """
        data = []
        for apt in apartments:
            # Получаем реакцию пользователя на эту квартиру
            user_reaction = await ReactionService.get_user_reaction(user_id, apt.id)
            reaction_emoji = ""
            if user_reaction == "like":
                reaction_emoji = " ❤️"
            elif user_reaction == "dislike":
                reaction_emoji = " 👎"
            
            # Собираем информацию о станциях метро
            metro_info = []
            for metro in apt.metro_stations:
                metro_info.append(f"{metro.station_name} ({metro.travel_time})")
            metro_str = "; ".join(metro_info) if metro_info else "Не указано"
            
            # Форматируем цены
            price_formatted = f"{apt.price:,} ₽" if apt.price else "Не указана"
            price_per_sqm_formatted = f"{apt.price_per_sqm:,} ₽/м²" if apt.price_per_sqm else "Не указана"
            
            data.append({
                'ID': apt.cian_id,
                'Название': apt.title + reaction_emoji,
                'Цена': price_formatted,
                'Цена за м²': price_per_sqm_formatted,
                'Цена (число)': apt.price or 0,
                'Цена за м² (число)': apt.price_per_sqm or 0,
                'Адрес': apt.address or "Не указан",
                'Станции метро': metro_str,
                'Ссылка': apt.url,
                'Дата добавления': apt.first_seen.strftime("%d.%m.%Y %H:%M"),
                'Последнее обновление': apt.last_updated.strftime("%d.%m.%Y %H:%M"),
                'Активно': "Да" if apt.is_active else "Нет"
            })
        
        return data
    
    @staticmethod
    def _create_formatted_excel(df: pd.DataFrame, filename: str, 
                               sheet_title: str = "Объявления о квартирах",
                               info_title: str = "Отчет по объявлениям о квартирах",
                               description: str = "Данные о квартирах"):
        """Создает красиво отформатированный Excel файл"""
        
        # Создаем новую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        
        # Добавляем данные
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Стили
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Arial', size=11)
        data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Форматируем заголовки
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Форматируем данные
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
        
        # Устанавливаем ширину колонок
        column_widths = {
            'A': 15,  # ID
            'B': 40,  # Название
            'C': 15,  # Цена
            'D': 15,  # Цена за м²
            'E': 12,  # Цена (число)
            'F': 15,  # Цена за м² (число)
            'G': 30,  # Адрес
            'H': 25,  # Станции метро
            'I': 15,  # Ссылка (скрыта)
            'J': 18,  # Дата добавления
            'K': 18,  # Последнее обновление
            'L': 10   # Активно
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Скрываем колонки с числовыми значениями цен (для сортировки)
        ws.column_dimensions['E'].hidden = True
        ws.column_dimensions['F'].hidden = True
        
        # Добавляем автофильтр
        ws.auto_filter.ref = f"A1:L{ws.max_row}"
        
        # Замораживаем первую строку
        ws.freeze_panes = 'A2'
        
        # Добавляем информационный лист
        info_ws = wb.create_sheet("Информация")
        ExcelExporter._add_info_sheet(info_ws, len(df), info_title, description)
        
        # Сохраняем файл
        wb.save(filename)
    
    @staticmethod
    def _add_info_sheet(ws, total_records: int, title: str = "Отчет по объявлениям о квартирах", 
                       description: str = "Данные о квартирах"):
        """Добавляет информационный лист с описанием данных"""
        
        info_data = [
            [title, ""],
            ["", ""],
            ["Дата создания:", datetime.now().strftime("%d.%m.%Y %H:%M")],
            ["Всего записей:", total_records],
            ["Источник данных:", "Cian.ru"],
            ["", ""],
            ["Описание колонок:", ""],
            ["ID", "Уникальный идентификатор объявления на Cian"],
            ["Название", "Заголовок объявления"],
            ["Цена", "Стоимость квартиры (отформатированная)"],
            ["Цена за м²", "Стоимость за квадратный метр (отформатированная)"],
            ["Адрес", "Адрес объекта недвижимости"],
            ["Станции метро", "Ближайшие станции метро с временем до них"],
            ["Дата добавления", "Когда объявление было впервые найдено"],
            ["Последнее обновление", "Когда информация об объявлении обновлялась"],
            ["Активно", "Доступно ли объявление на данный момент"],
            ["", ""],
            ["Примечания:", ""],
            ["• Данные отсортированы по возрастанию цены", ""],
            ["• Используйте автофильтр для поиска и сортировки", ""],
            ["• Скрытые колонки содержат числовые значения для сортировки", ""],
        ]
        
        for row_data in info_data:
            ws.append(row_data)
        
        # Форматирование информационного листа
        title_font = Font(name='Arial', size=16, bold=True, color='366092')
        header_font = Font(name='Arial', size=12, bold=True)
        data_font = Font(name='Arial', size=11)
        
        ws['A1'].font = title_font
        
        for row in ws.iter_rows(min_row=3, max_row=6):
            row[0].font = header_font
        
        for row in ws.iter_rows(min_row=8, max_row=16):
            row[0].font = header_font
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50

    @staticmethod
    async def export_statistics_to_excel(filename: Optional[str] = None) -> str:
        """
        Экспортирует статистику по базе данных в Excel
        
        Args:
            filename: Имя файла
            
        Returns:
            str: Путь к созданному файлу
        """
        
        # Получаем статистику
        stats = await ApartmentService.get_statistics()
        apartments = await ApartmentService.get_apartments(only_active=True)
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"statistics_export_{timestamp}.xlsx"
        
        # Создаем книгу Excel
        wb = Workbook()
        
        # Лист со статистикой
        stats_ws = wb.active
        stats_ws.title = "Общая статистика"
        
        stats_data = [
            ["Показатель", "Значение"],
            ["Всего объявлений", stats['total_apartments']],
            ["Активных объявлений", stats['active_apartments']],
            ["Неактивных объявлений", stats['inactive_apartments']],
            ["Средняя цена", f"{stats['average_price']:,} ₽"],
            ["Дата создания отчета", datetime.now().strftime("%d.%m.%Y %H:%M")]
        ]
        
        for row in stats_data:
            stats_ws.append(row)
        
        # Форматирование
        for cell in stats_ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        stats_ws.column_dimensions['A'].width = 25
        stats_ws.column_dimensions['B'].width = 20
        
        # Лист с топ дешевых квартир
        if apartments:
            cheap_ws = wb.create_sheet("Топ дешевых квартир")
            cheap_apartments = sorted(apartments, key=lambda x: x.price or float('inf'))[:20]
            
            cheap_data = [["Место", "Цена", "Название", "Станции метро"]]
            for i, apt in enumerate(cheap_apartments, 1):
                metro_info = ", ".join([m.station_name for m in apt.metro_stations[:2]])
                cheap_data.append([
                    i,
                    f"{apt.price:,} ₽" if apt.price else "Не указана",
                    apt.title,
                    metro_info or "Не указано"
                ])
            
            for row in cheap_data:
                cheap_ws.append(row)
            
            # Форматирование заголовков
            for cell in cheap_ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFF')
            
            cheap_ws.column_dimensions['A'].width = 8
            cheap_ws.column_dimensions['B'].width = 15
            cheap_ws.column_dimensions['C'].width = 40
            cheap_ws.column_dimensions['D'].width = 25
        
        wb.save(filename)
        return os.path.abspath(filename)